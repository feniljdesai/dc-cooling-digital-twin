import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import streamlit as st
from io import BytesIO
from openpyxl import Workbook

# ---------------------------
# Simple Cooling Plant Dashboard
# ---------------------------
# Not a full ASHRAE compliance engine yet.
# It's a sizing + transient-response sandbox with basic redundancy + BOM outputs.

WATER_CP = 4180.0  # J/kg-K
WATER_RHO = 997.0  # kg/m3

def cop_aircooled(T_amb_C, cop_ref=6.0, T_ref=25.0, k=0.08, cop_min=2.5):
    cop = cop_ref - k * (T_amb_C - T_ref)
    return max(cop, cop_min)

def simulate(
    dt_s,
    duration_min,
    T_set_C,
    C_th_J_per_K,
    Q_cool_max_MW,
    Kp,
    T_amb_C,
    Q_it_MW,
    cop_ref,
    T_ref,
    k_cop,
    cop_min,
):
    n = int(duration_min * 60 / dt_s)
    t_s = np.arange(n) * dt_s
    t_min = t_s / 60.0

    # profiles
    T_amb = np.ones(n) * T_amb_C
    Q_it = np.ones(n) * (Q_it_MW * 1e6)

    # states
    T_supply = np.zeros(n)
    T_supply[0] = T_set_C
    Q_cool = np.zeros(n)
    P_ch = np.zeros(n)

    Q_cool_max_W = Q_cool_max_MW * 1e6

    for i in range(1, n):
        e = T_supply[i-1] - T_set_C
        u = np.clip(Kp * e, 0.0, 1.0)

        Q_cool[i] = u * Q_cool_max_W
        dT = (Q_it[i] - Q_cool[i]) * dt_s / C_th_J_per_K
        T_supply[i] = T_supply[i-1] + dT

        cop = cop_aircooled(T_amb[i], cop_ref=cop_ref, T_ref=T_ref, k=k_cop, cop_min=cop_min)
        P_ch[i] = Q_cool[i] / cop

    # KPIs
    max_excursion_C = float(np.max(T_supply - T_set_C))
    peak_ch_MW = float(np.max(P_ch) / 1e6)

    # manual trapezoid energy
    energy_Wh = 0.0
    for i in range(1, n):
        energy_Wh += 0.5 * (P_ch[i] + P_ch[i-1]) * dt_s / 3600.0
    energy_MWh = float(energy_Wh / 1e6)

    df = pd.DataFrame({
        "time_min": t_min,
        "T_supply_C": T_supply,
        "Q_it_MW": Q_it / 1e6,
        "Q_cool_MW": Q_cool / 1e6,
        "P_chiller_MW": P_ch / 1e6,
        "T_amb_C": T_amb
    })

    return df, {
        "max_supply_temp_excursion_C": max_excursion_C,
        "peak_chiller_power_MW": peak_ch_MW,
        "chiller_energy_MWh": energy_MWh
    }

def size_hydronics(Q_MW, deltaT_C, pump_eff=0.75, dp_bar=1.5):
    """
    Very simple hydronic sizing:
      m_dot = Q / (cp * dT)
      V_dot = m_dot / rho
      Pump power = dp * Vdot / eff
    dp_bar is total loop DP in bar (rough placeholder).
    """
    Q_W = Q_MW * 1e6
    m_dot = Q_W / (WATER_CP * max(deltaT_C, 0.1))  # kg/s
    V_dot_m3_s = m_dot / WATER_RHO
    dp_Pa = dp_bar * 1e5
    P_pump_W = dp_Pa * V_dot_m3_s / max(pump_eff, 0.1)
    return m_dot, V_dot_m3_s, P_pump_W

def redundancy_counts(required_units, redundancy_mode):
    """
    required_units = N
    modes:
      N: N
      N+1: N+1
      2N: 2N
    """
    if redundancy_mode == "N":
        return required_units
    if redundancy_mode == "N+1":
        return required_units + 1
    if redundancy_mode == "2N":
        return required_units * 2
    return required_units

def make_excel(df, kpis, bom_df):
    wb = Workbook()
    ws = wb.active
    ws.title = "timeseries"
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))

    ws2 = wb.create_sheet("kpis")
    ws2.append(["metric", "value"])
    for k, v in kpis.items():
        ws2.append([k, float(v)])

    ws3 = wb.create_sheet("BOM")
    ws3.append(list(bom_df.columns))
    for row in bom_df.itertuples(index=False):
        ws3.append(list(row))

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

# ---------------------------
# UI
# ---------------------------
st.set_page_config(page_title="DC Cooling Sizing Dashboard", layout="wide")
st.title("Data Center Cooling Sizing Dashboard (beginner model)")

with st.sidebar:
    st.header("Inputs")

    IT_MW = st.slider("IT Load (MW)", 1.0, 200.0, 8.0, 1.0)
    cooling_overhead = st.slider("Facility-to-IT cooling overhead factor", 1.0, 1.4, 1.1, 0.01)
    Q_required_MW = IT_MW * cooling_overhead

    st.caption("This sets required heat removal. (Later we can model PUE/90.4 properly.)")

    T_supply_set = st.slider("CHW Supply Setpoint (°C)", 4.0, 20.0, 7.0, 0.5)
    deltaT = st.slider("CHW ΔT (°C)", 2.0, 12.0, 5.0, 0.5)
    T_amb = st.slider("Ambient Dry-bulb (°C)", -5.0, 45.0, 32.0, 0.5)

    st.divider()
    st.subheader("Chillers")
    ch_unit_MW = st.slider("Chiller unit size (MW)", 0.5, 20.0, 5.0, 0.5)
    redundancy_mode = st.selectbox("Redundancy", ["N", "N+1", "2N"])

    cop_ref = st.slider("COP @ Tref", 2.0, 10.0, 6.0, 0.1)
    T_ref = st.slider("Tref (°C)", 10.0, 35.0, 25.0, 0.5)
    k_cop = st.slider("COP drop per °C", 0.00, 0.20, 0.08, 0.01)
    cop_min = st.slider("Minimum COP", 1.5, 5.0, 2.5, 0.1)

    st.divider()
    st.subheader("Dynamics")
    dt_s = st.selectbox("Timestep (s)", [5, 10, 30, 60], index=1)
    duration_min = st.selectbox("Sim duration (min)", [30, 60, 120, 240], index=2)
    C_th = st.slider("Thermal mass C_th (J/K)", 5e7, 8e8, 2.5e8, 5e7)
    Kp = st.slider("Control gain Kp", 0.01, 1.00, 0.15, 0.01)

    st.divider()
    st.subheader("Pumps (rough)")
    dp_bar = st.slider("Total loop DP (bar)", 0.2, 4.0, 1.5, 0.1)
    pump_eff = st.slider("Pump efficiency", 0.3, 0.9, 0.75, 0.05)

# --- Sizing ---
N_required = int(np.ceil(Q_required_MW / ch_unit_MW))
units_installed = redundancy_counts(N_required, redundancy_mode)
installed_capacity_MW = units_installed * ch_unit_MW

m_dot, V_dot, P_pump_W = size_hydronics(Q_required_MW, deltaT, pump_eff=pump_eff, dp_bar=dp_bar)

bom = [
    ["Chiller", units_installed, f"{ch_unit_MW:.1f} MW each", "Air-cooled (placeholder model)"],
    ["Primary CHW Pump", 2 if redundancy_mode != "N" else 1, f"{V_dot*3600:.0f} m3/h each", "Assumed parallel pumps"],
    ["CHW Supply/Return Headers", 1, "-", "Conceptual"],
    ["Controls + Sensors", 1, "-", "Temp sensors, control valve logic (conceptual)"],
]
bom_df = pd.DataFrame(bom, columns=["Item", "Qty", "Size/Rating", "Notes"])

# --- Sim (use installed capacity as max) ---
df, kpis = simulate(
    dt_s=dt_s,
    duration_min=duration_min,
    T_set_C=T_supply_set,
    C_th_J_per_K=C_th,
    Q_cool_max_MW=installed_capacity_MW,
    Kp=Kp,
    T_amb_C=T_amb,
    Q_it_MW=Q_required_MW,
    cop_ref=cop_ref,
    T_ref=T_ref,
    k_cop=k_cop,
    cop_min=cop_min,
)

# --- Layout ---
col1, col2 = st.columns([1.2, 1])

with col1:
    st.subheader("Plots")

    fig1 = plt.figure()
    plt.plot(df["time_min"], df["Q_it_MW"], label="Required Cooling (MW)")
    plt.plot(df["time_min"], df["Q_cool_MW"], label="Cooling Delivered (MW)")
    plt.xlabel("Time (min)")
    plt.ylabel("MW")
    plt.title("Load vs Cooling")
    plt.legend()
    st.pyplot(fig1, clear_figure=True)

    fig2 = plt.figure()
    plt.plot(df["time_min"], df["T_supply_C"], label="Supply Temp (°C)")
    plt.axhline(T_supply_set, linestyle="--", label="Setpoint")
    plt.xlabel("Time (min)")
    plt.ylabel("°C")
    plt.title("Supply Temperature Response")
    plt.legend()
    st.pyplot(fig2, clear_figure=True)

    fig3 = plt.figure()
    plt.plot(df["time_min"], df["P_chiller_MW"], label="Chiller Electric Power (MW)")
    plt.xlabel("Time (min)")
    plt.ylabel("MW")
    plt.title("Chiller Power")
    plt.legend()
    st.pyplot(fig3, clear_figure=True)

with col2:
    st.subheader("Sizing Summary")

    st.metric("Required Cooling (MW)", f"{Q_required_MW:.1f}")
    st.metric("Installed Capacity (MW)", f"{installed_capacity_MW:.1f}")
    st.metric("Chiller Units (installed)", f"{units_installed} ({redundancy_mode})")
    st.metric("Flow (kg/s)", f"{m_dot:.1f}")
    st.metric("Flow (m³/h)", f"{V_dot*3600:.0f}")
    st.metric("Pump Power (kW, rough)", f"{P_pump_W/1000:.1f}")

    st.divider()
    st.subheader("KPIs")
    for k, v in kpis.items():
        st.write(f"**{k}**: {v:.3f}")

    st.divider()
    st.subheader("BOM")
    st.dataframe(bom_df, use_container_width=True)

    st.divider()
    excel_bytes = make_excel(df, kpis, bom_df)
    st.download_button(
        label="Download Excel (timeseries + KPIs + BOM)",
        data=excel_bytes,
        file_name="dc_cooling_results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

st.subheader("Time Series Table")
st.dataframe(df.head(200), use_container_width=True)